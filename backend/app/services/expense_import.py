import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import Allocation, AuditLog, Counterparty, Expense, Invoice, Partner, Payment, Store, Tag
from app.services.finance import distribute_evenly


def _header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


HEADERS = {
    "partner": {"наименование контрагента"},
    "service": {"суть рекламного сообщения"},
    "tag": {"свид рекламы", "вид рекламы"},
    "period": {"месяц/год оказания услуг"},
    "invoice_number": {"№ счета", "номер счета"},
    "invoice_date": {"дата счета"},
    "amount": {"сумма счета с ндс"},
}


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _invoice_number(value: object) -> str:
    """Normalize cash and missing invoice numbers used in spreadsheets."""
    number = _text(value)
    if number.casefold() == "нал":
        return "Наличные"
    return number or "б/н"


def _decimal(value: object) -> Decimal:
    if isinstance(value, (Decimal, int, float)):
        result = Decimal(str(value))
    else:
        cleaned = re.sub(r"[^\d,.-]", "", _text(value).replace("\xa0", "").replace(" ", "")).strip(".")
        if cleaned.count(",") == 1 and "." not in cleaned:
            cleaned = cleaned.replace(",", ".")
        try:
            result = Decimal(cleaned)
        except InvalidOperation as error:
            raise ValueError(f"некорректная сумма «{value}»") from error
    if not result.is_finite():
        raise ValueError("сумма должна быть конечным числом")
    return result


def _date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(_text(value), fmt).date()
        except ValueError:
            pass
    raise ValueError(f"некорректная дата счета «{_text(value)}»")


MONTHS = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
    "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}


def _period(value: object) -> tuple[int, int]:
    if isinstance(value, (datetime, date)):
        return value.month, value.year
    text = _text(value).casefold()
    year_match = re.search(r"\b(20\d{2}|21\d{2})\b", text)
    month = next((number for name, number in MONTHS.items() if name in text), None)
    numeric = re.search(r"\b(0?[1-9]|1[0-2])[./-](20\d{2}|21\d{2})\b", text)
    if numeric:
        return int(numeric.group(1)), int(numeric.group(2))
    if month and year_match:
        return month, int(year_match.group(1))
    raise ValueError(f"некорректный месяц/год «{_text(value)}»")


def _find_header(rows) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(rows, 1):
        normalized = [_header(value) for value in row]
        columns = {field: next((i for i, value in enumerate(normalized) if value in aliases), -1) for field, aliases in HEADERS.items()}
        if all(index >= 0 for index in columns.values()):
            return row_number, columns
    raise ValueError("не найдена строка заголовков или отсутствуют обязательные колонки")


def _read_rows(content: bytes, filename: str) -> list[tuple[object, ...]]:
    if filename.casefold().endswith(".xls"):
        import xlrd

        try:
            workbook = xlrd.open_workbook(file_contents=content)
        except Exception as error:
            raise ValueError("файл не является корректной книгой XLS") from error
        sheet = workbook.sheet_by_index(0)
        rows = []
        for row_index in range(sheet.nrows):
            values = []
            for cell in sheet.row(row_index):
                if cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
                elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    values.append(None)
                else:
                    values.append(cell.value)
            rows.append(tuple(values))
        return rows
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as error:
        raise ValueError("файл не является корректной книгой XLSX") from error
    sheet = workbook.active
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def import_expenses_excel(content: bytes, filename: str, db: Session) -> dict:
    rows = _read_rows(content, filename)
    header_row, columns = _find_header(rows)
    header_values = rows[header_row - 1]
    stores = db.scalars(select(Store).where(Store.is_active.is_(True))).all()
    stores_by_name = {_header(store.name): store for store in stores}
    store_columns = [(i, stores_by_name[name]) for i, value in enumerate(header_values) if (name := _header(value)) in stores_by_name]
    loaded = 0
    errors: list[dict] = []

    for row_number, values in enumerate(rows[header_row:], header_row + 1):
        if not any(value not in (None, "") for value in values):
            continue
        try:
            def required(field: str, label: str) -> object:
                value = values[columns[field]] if columns[field] < len(values) else None
                if value is None or not _text(value):
                    raise ValueError(f"не заполнено поле «{label}»")
                return value

            month, year = _period(required("period", "месяц/год оказания услуг"))
            amount = _decimal(required("amount", "Сумма счета с НДС"))
            if amount < 0:
                raise ValueError("сумма счета не может быть отрицательной")
            allocations = []
            for index, store in store_columns:
                value = values[index] if index < len(values) else None
                if value not in (None, "") and (allocation := _decimal(value)) > 0:
                    allocations.append((store, allocation))
            partner_name = _text(required("partner", "Наименование контрагента"))
            service = _text(required("service", "Суть рекламного сообщения"))
            tag_name = _text(values[columns["tag"]]) or None
            invoice_number_value = values[columns["invoice_number"]] if columns["invoice_number"] < len(values) else None
            invoice_number = _invoice_number(invoice_number_value)
            invoice_date = _date(required("invoice_date", "Дата счета"))

            with db.begin_nested():
                partner = db.scalar(select(Partner).where(Partner.deleted_at.is_(None), func.lower(Partner.name) == partner_name.lower()))
                if not partner:
                    partner = Partner(name=partner_name, comment=None); db.add(partner); db.flush()
                counterparty = db.scalar(select(Counterparty).where(Counterparty.deleted_at.is_(None), Counterparty.partner_id == partner.id, func.lower(Counterparty.full_name) == partner_name.lower()))
                if not counterparty:
                    counterparty = Counterparty(partner_id=partner.id, full_name=partner_name, short_name=None, entity_type="organization", inn=None, kpp=None, comment=None); db.add(counterparty); db.flush()
                expense = Expense(partner_id=partner.id, counterparty_id=counterparty.id, service_name=service, contract_number=None, contract_date=None, expense_month=month, expense_year=year, comment=None)
                if tag_name:
                    tag = db.scalar(select(Tag).where(func.lower(Tag.name) == tag_name.lower()))
                    if not tag:
                        tag = Tag(name=tag_name); db.add(tag); db.flush()
                    expense.tags = [tag]
                db.add(expense); db.flush()
                invoice = Invoice(expense_id=expense.id, invoice_number=invoice_number, invoice_date=invoice_date, amount=amount, vat_amount=None, comment=None)
                db.add(invoice); db.flush()
                db.add(Payment(invoice_id=invoice.id, payment_date=invoice_date, amount=amount, comment="Импортировано из Excel"))
                distributed = distribute_evenly(amount, len(allocations))
                for (store, _), allocation_amount in zip(allocations, distributed):
                    db.add(Allocation(expense_id=expense.id, store_id=store.id, amount=allocation_amount))
                db.add(AuditLog(entity_type="expense", entity_id=expense.id, action="imported", metadata_={"source": "xlsx", "row": row_number}, created_at=datetime.now(timezone.utc)))
            loaded += 1
        except Exception as error:
            errors.append({"row": row_number, "message": str(error) or error.__class__.__name__})
    db.commit()
    return {"loaded": loaded, "errors_count": len(errors), "errors": errors}
