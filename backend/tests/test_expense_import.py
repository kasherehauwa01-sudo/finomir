from datetime import date
from decimal import Decimal
from io import BytesIO
from contextlib import nullcontext
import uuid

import pytest
from openpyxl import Workbook

from app.models import Allocation, Invoice, Payment, Store
from app.services.expense_import import _date, _decimal, _find_header, _invoice_date, _invoice_number, _period, import_expenses_excel


class Result:
    def __init__(self, items): self.items = items
    def all(self): return self.items


class ImportDb:
    def __init__(self, stores): self.stores = stores; self.added = []
    def scalars(self, query): return Result(self.stores if query.column_descriptions[0].get("entity") is Store else [])
    def scalar(self, _): return None
    def add(self, item): self.added.append(item)
    def flush(self):
        for item in self.added:
            if hasattr(item, "id") and item.id is None: item.id = uuid.uuid4()
    def begin_nested(self): return nullcontext()
    def commit(self): pass


def test_import_value_formats():
    assert _period("Февраль 2016") == (2, 2016)
    assert _period("03/2026") == (3, 2026)
    assert _date("21.01.2016") == date(2016, 1, 21)
    assert _decimal("34 000,00р.") == Decimal("34000.00")
    assert _invoice_number("нал") == "Наличные"
    assert _invoice_number(" НАЛ ") == "Наличные"
    assert _invoice_number("15") == "15"
    assert _invoice_number(None) == "б/н"
    assert _invoice_number("  ") == "б/н"


@pytest.mark.parametrize(("source", "expected"), [
    ("21.06.16.", date(2016, 6, 21)),
    ("29,06.2016", date(2016, 6, 29)),
    ("24.10.216", date(2016, 10, 24)),
    ("12,07,17", date(2017, 7, 12)),
    ("03.09.19", date(2019, 9, 3)),
    ("10.07,2024", date(2024, 7, 10)),
])
def test_invoice_date_formats_are_normalized(source, expected):
    assert _date(source) == expected


@pytest.mark.parametrize("source", [None, "", "дата неизвестна", "32.18.2024"])
def test_invalid_invoice_date_uses_expense_period(source):
    assert _invoice_date(source, 2, 2026) == date(2026, 2, 1)


def test_header_row_can_contain_newlines_and_typo_from_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Отчет", None])
    sheet.append([
        "Наименование контрагента", "Суть рекламного сообщения\n", "Свид рекламы",
        "месяц/год оказания услуг ", "№ счета\n", "Дата счета", "Сумма счета с НДС\n",
    ])
    row, columns = _find_header(tuple(tuple(cell.value for cell in row) for row in sheet.iter_rows()))
    assert row == 2
    assert set(columns) == {"partner", "service", "tag", "period", "invoice_number", "invoice_date", "amount"}


def test_invalid_period_has_readable_error():
    with pytest.raises(ValueError, match="некорректный месяц/год"):
        _period("когда-нибудь")


def test_excel_import_creates_payment_and_preserves_store_amounts():
    workbook = Workbook(); sheet = workbook.active
    sheet.append(["Наименование контрагента", "Суть рекламного сообщения", "Вид рекламы", "месяц/год оказания услуг", "№ счета", "Дата счета", "Сумма счета с НДС", "Магазин 1", "Магазин 2"])
    sheet.append(["Партнер", "Реклама", "Интернет", "Январь 2026", "15", "10.01.2026", 100, "70,25", "15,50"])
    content = BytesIO(); workbook.save(content)
    stores = [Store(id=uuid.uuid4(), name=f"Магазин {number}", address=None, comment=None, is_active=True, is_system=False) for number in (1, 2)]
    db = ImportDb(stores)

    result = import_expenses_excel(content.getvalue(), "expenses.xlsx", db)

    payment = next(item for item in db.added if isinstance(item, Payment))
    allocations = [item.amount for item in db.added if isinstance(item, Allocation)]
    assert result == {"loaded": 1, "errors_count": 0, "errors": []}
    assert payment.amount == Decimal("100")
    assert allocations == [Decimal("70.25"), Decimal("15.50")]


def test_excel_import_marks_nal_invoice_as_cash():
    workbook = Workbook(); sheet = workbook.active
    sheet.append(["Наименование контрагента", "Суть рекламного сообщения", "Вид рекламы", "месяц/год оказания услуг", "№ счета", "Дата счета", "Сумма счета с НДС"])
    sheet.append(["Партнер", "Реклама", "Интернет", "Январь 2026", "нал", "10.01.2026", 100])
    content = BytesIO(); workbook.save(content)
    db = ImportDb([])

    result = import_expenses_excel(content.getvalue(), "expenses.xlsx", db)

    invoice = next(item for item in db.added if isinstance(item, Invoice))
    assert result == {"loaded": 1, "errors_count": 0, "errors": []}
    assert invoice.invoice_number == "Наличные"


def test_excel_import_uses_default_when_invoice_number_is_missing():
    workbook = Workbook(); sheet = workbook.active
    sheet.append(["Наименование контрагента", "Суть рекламного сообщения", "Вид рекламы", "месяц/год оказания услуг", "№ счета", "Дата счета", "Сумма счета с НДС"])
    sheet.append(["Партнер", "Реклама", "Интернет", "Январь 2026", None, "10.01.2026", 100])
    content = BytesIO(); workbook.save(content)
    db = ImportDb([])

    result = import_expenses_excel(content.getvalue(), "expenses.xlsx", db)

    invoice = next(item for item in db.added if isinstance(item, Invoice))
    assert result == {"loaded": 1, "errors_count": 0, "errors": []}
    assert invoice.invoice_number == "б/н"


def test_excel_import_uses_period_start_when_invoice_date_is_missing():
    workbook = Workbook(); sheet = workbook.active
    sheet.append(["Наименование контрагента", "Суть рекламного сообщения", "Вид рекламы", "месяц/год оказания услуг", "№ счета", "Дата счета", "Сумма счета с НДС"])
    sheet.append(["Партнер", "Реклама", "Интернет", "Февраль 2026", "15", None, 100])
    content = BytesIO(); workbook.save(content)
    db = ImportDb([])

    result = import_expenses_excel(content.getvalue(), "expenses.xlsx", db)

    invoice = next(item for item in db.added if isinstance(item, Invoice))
    payment = next(item for item in db.added if isinstance(item, Payment))
    assert result == {"loaded": 1, "errors_count": 0, "errors": []}
    assert invoice.invoice_date == date(2026, 2, 1)
    assert payment.payment_date == date(2026, 2, 1)
